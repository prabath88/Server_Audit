import subprocess
import os
import sys

sar_path=""

try:
    msg=subprocess.check_output("sudo yum install python36u -y",shell=True);
    msg=subprocess.check_output("sudo yum -y install python36u-pip",shell=True);
    msg=subprocess.check_output("sudo pip3.6 install xlsxwriter",shell=True);
    msg=subprocess.check_output("sudo pip3.6 install pandas",shell=True);
    msg=subprocess.check_output("sudo pip3.6 install openpyxl",shell=True);
    sar_path="rpm" 
except:
    print ("Error ...! installing python3 , setup continue with different settings")
    sar_path="deb"
    

if sar_path == 'deb':
 try:
   msg=subprocess.check_output("sudo apt-get install python3.5 -y",shell=True);
   msg=subprocess.check_output("sudo apt-get install -y python3-pip",shell=True);
   msg=subprocess.check_output("sudo pip3 install xlsxwriter",shell=True);
   msg=subprocess.check_output("sudo pip3 install pandas",shell=True);
   msg=subprocess.check_output("sudo pip3 install openpyxl",shell=True);
 except:
   print ("Error..! installing python3, setup continue.. please install xlsxwriter,pandas and openpyxl manually" )
   sar_path="rpm"
   



import xlsxwriter
import pandas as pd
import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
import datetime 
import time


if sar_path == 'deb':
       bashscript=subprocess.check_output("sh ./init/user-audit_deb.sh ",shell=True)
elif sar_path == 'rpm':
       bashscript=subprocess.check_output("sh ./init/user-audit_rpm.sh ",shell=True)
else:
       print ("ll")

wb = openpyxl.load_workbook("server.xlsx")
ws = wb.active

        
infile = open('avg', 'r')
cpu_numbers = [float(line) for line in infile.readlines()]
infile.close()
cpu_mean = sum(cpu_numbers)/len(cpu_numbers)
#print(cpu_mean)


infile = open('mem', 'r')
mem_numbers = [float(line) for line in infile.readlines()]
infile.close()
mem_mean = sum(mem_numbers)/len(mem_numbers)
#print(mem_mean)


mon = datetime.datetime.now().strftime("%m")
month = str(mon)

if month == '01':
   raw=str(15)
elif month == '02':
   raw=str(16)
elif month == '03':
   raw=str(17)
elif month == '04':
   raw=str(18)
elif month == '05':
   raw=str(19)
elif month == '06':
   raw=str(20)
elif month == '07':
   raw=str(21)
elif month == '08':
   raw=str(22)
elif month == '09':
   raw=str(23)
elif month == '10':
   raw=str(24)
elif month == '11':
   raw=str(25)
else:
   raw=str(26)


cpucell="B"+raw
memcell="C"+raw

ws[cpucell] = str(cpu_mean)
ws[memcell] = str(mem_mean)
#----------------------------------------------------------------------------------------------------------------
root_total=subprocess.check_output("df -h | grep -sw '/' | awk '{print $2}' | awk -FG '{print $1}'",shell=True)
root_t = float(root_total)


root_use=subprocess.check_output("df -h | grep -sw '/' | awk '{print $2}' | awk -FG '{ print $1}'",shell=True)
root_u = float(root_use)



boot_total=subprocess.check_output("df -h | grep -sw '/boot' | awk '{print $2}' | awk -FM '{print $1}'",shell=True)
boot_t = float(boot_total)/1024


boot_use=subprocess.check_output("df -h | grep -sw '/boot' | awk '{print $3}' | awk -FM '{print $1}'",shell=True)
boot_u = float(boot_use)/1024



if month == '01':
   raw=str(37)
elif month == '02':
   raw=str(38)
elif month == '03':
   raw=str(39)
elif month == '04':
   raw=str(40)
elif month == '05':
   raw=str(41)
elif month == '06':
   raw=str(42)
elif month == '07':
   raw=str(43)
elif month == '08':
   raw=str(44)
elif month == '09':
   raw=str(45)
elif month == '10':
   raw=str(46)
elif month == '11':
   raw=str(47)
else:
   raw=str(48)




boot_t_cell="B"+raw
boot_u_cell="C"+raw
root_t_cell="D"+raw
root_u_cell="E"+raw


ws[boot_t_cell] = str(boot_t)
ws[boot_u_cell] = str(boot_u)
ws[root_t_cell] = str(root_t)
ws[root_u_cell] = str(root_u)

#==========================================
infile = open('Employees', 'r')
emp = [str(line) for line in infile.readlines()]
infile.close()


count = 1
emp_raw=54
while (count < len(emp)):
    emp_cell="A"+str(emp_raw)
    ws[emp_cell]= str(emp[count])
    count = count + 1
    emp_raw = emp_raw + 1
#============================================

infile = open('Usernames', 'r')
usr = [str(line) for line in infile.readlines()]
infile.close()

count = 1
usr_raw=54
while (count < len(usr)):
    usr_cell="B"+str(usr_raw)
    ws[usr_cell]= str(usr[count])
    count = count + 1
    usr_raw = usr_raw + 1

#================================================

infile = open('User-groups', 'r')
usr_grp = [str(line) for line in infile.readlines()]
infile.close()

count = 1
usr_grp_raw=54
while (count < len(usr_grp)):
    usr_grp_cell="D"+str(usr_grp_raw)
    ws[usr_grp_cell]= str(usr_grp[count])
    count = count + 1
    usr_grp_raw = usr_grp_raw + 1

usr_grp_raw=54
count = 1
while (count < len(usr_grp)):
    usr_grp_cell="E"+str(usr_grp_raw)
    ws[usr_grp_cell]=str(time.strftime("%Y-%m-%d"))
    count = count + 1
    usr_grp_raw = usr_grp_raw + 1
#================================================
infile = open('Groups', 'r')
grp = [str(line) for line in infile.readlines()]
infile.close()

infile = open('Members', 'r')
mem = [str(line) for line in infile.readlines()]
infile.close()

raw=120
count = 0
while (count < len(grp)):
    grp_cell="A"+str(raw)
    #mem_cell="B"+str(raw)
    date_cell="C"+str(raw)
    ws[date_cell]=str(time.strftime("%Y-%m-%d"))
    #ws[mem_cell]= str(mem[count])
    ws[grp_cell]= str(grp[count])
    count = count + 1
    raw = raw + 1

raw=120
count = 0
while (count < len(mem)):
    mem_cell="B"+str(raw)
    ws[mem_cell]= str(mem[count])
    count = count + 1
    raw = raw + 1





wb.save("server.xlsx")

clean=subprocess.check_output("rm avg  Employees  group-members-tmp  Groups  groups-tmp  mem  Members User-groups  Usernames -rf",shell=True)
